"""
evaluate_rag.py
===============
End-to-end evaluation harness for the NIFTY Graph RAG chatbot.

Metrics
-------
Per-question (LLM-as-judge, scored 1-5):
  - Faithfulness      - answer is grounded in retrieved context, no hallucinations
  - Answer Relevance  - answer directly addresses the question asked
  - Completeness      - all parts of the question are addressed

System:
  - Latency (s)
  - Context Hit       - did the RETRIEVER actually surface the expected entity in
                         its retrieved documents? (measures retrieval quality)
  - Answer Mentions Entity - does the final ANSWER text mention the expected entity?
                         (measures generation/grounding behavior, NOT retrieval —
                         renamed from the old "anchor_hit" which conflated the two)
  - Error rate

CHANGELOG (vs. previous version)
---------------------------------
- FIX: `res["answer"]` being None (seen with gpt-oss in prior runs) crashed the
  regex think-tag stripper with a silent TypeError caught by the outer except.
  Now explicitly checked and raised as a clear, loggable error.
- FIX: when `result.error` is set, the judge is now SKIPPED entirely instead of
  scoring a placeholder answer. Previously this could produce misleading 0-scores
  in the "avg_score" column that look like real (bad) judge scores rather than
  upstream failures. Failed rows are now unambiguous: no scores, error text visible.
- FIX: added retry-with-backoff on Groq 429 rate-limit errors instead of failing
  the question outright (this is what killed qwen's C01 in the last run).
- FIX: the blanket `sleep(6.0)` before every single question — regardless of
  provider — is now provider-aware. Ollama (local, no rate limit) doesn't need
  to be throttled; only Groq-backed models do, and the delay is configurable.
- NEW: added a real retrieval-quality metric ("context_hit") that inspects what
  the retriever actually returned, separate from whether the model's answer text
  happens to mention the entity. This is what "anchor_hit" was implicitly assumed
  to measure before, but didn't.
- NEW: comparison-category questions (multiple expected_entities) now also track
  whether ALL expected entities were hit, not just ANY — "any" was hiding partial
  answers in comparison questions as full hits.
- NEW: retrieved context is captured and written to the Excel output per-question
  for manual debugging, so you can see exactly what the retriever handed the LLM.

Usage
-----
python evaluate_rag.py                          # uses defaults (localhost:8000)
python evaluate_rag.py --dataset custom_qs.json
python evaluate_rag.py --out results.xlsx
python evaluate_rag.py --throttle-s 2.0          # seconds between Groq-backed calls
"""

import argparse
import asyncio
import json
import os
import re
import time
import uuid
from dataclasses import dataclass, field, asdict
from pathlib import Path
from typing import Optional

import httpx
from dotenv import load_dotenv
from openai import OpenAI          # lightweight judge - works with Ollama too
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel

os.environ.pop("SSL_CERT_FILE", None)
os.environ.pop("REQUESTS_CA_BUNDLE", None)
os.environ.pop("CURL_CA_BUNDLE", None)

load_dotenv()

# ─────────────────────────────────────────────────────────────────────────────
# DEFAULT NIFTY 50 EVALUATION DATASET
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_DATASET = [
    {"id": "F01", "category": "Single-Company Factual", "question": "What is the current revenue of TCS?", "expected_entities": ["TCS"], "ground_truth": None},
    {"id": "F02", "category": "Single-Company Factual", "question": "Who leads Reliance Industries?", "expected_entities": ["RELIANCE"], "ground_truth": None},
    {"id": "F03", "category": "Single-Company Factual", "question": "What sector does Infosys operate in?", "expected_entities": ["INFY"], "ground_truth": None},
    {"id": "F04", "category": "Single-Company Factual", "question": "What is HDFC Bank's net profit for the latest reported quarter?", "expected_entities": ["HDFCBANK"], "ground_truth": None},
    {"id": "F05", "category": "Single-Company Factual", "question": "What is the debt-to-equity ratio of Tata Motors?", "expected_entities": ["TATAMOTORS"], "ground_truth": None},
    {"id": "C01", "category": "Comparison", "question": "Compare the revenue of TCS and Infosys.", "expected_entities": ["TCS", "INFY"], "ground_truth": None},
    {"id": "C02", "category": "Comparison", "question": "Which has a higher market cap, Wipro or HCL Technologies?", "expected_entities": ["WIPRO", "HCLTECH"], "ground_truth": None},
    {"id": "C03", "category": "Comparison", "question": "How does HDFC Bank's PAT compare to ICICI Bank?", "expected_entities": ["HDFCBANK", "ICICIBANK"], "ground_truth": None},
    {"id": "S01", "category": "Sector", "question": "Which NIFTY 50 companies belong to the IT sector?", "expected_entities": ["IT"], "ground_truth": None},
    {"id": "S02", "category": "Sector", "question": "List all banking companies in the NIFTY 50.", "expected_entities": ["Banking", "Financial Services"], "ground_truth": None},
    {"id": "S03", "category": "Sector", "question": "Which energy sector stocks are part of NIFTY 50?", "expected_entities": ["Energy", "Oil"], "ground_truth": None},
    {"id": "M01", "category": "Financial Metric", "question": "What is the P/E ratio of Bajaj Finance?", "expected_entities": ["BAJFINANCE"], "ground_truth": None},
    {"id": "M02", "category": "Financial Metric", "question": "What is the EPS of Asian Paints?", "expected_entities": ["ASIANPAINT"], "ground_truth": None},
    {"id": "M03", "category": "Financial Metric", "question": "What is the dividend yield of Coal India?", "expected_entities": ["COALINDIA"], "ground_truth": None},
    {"id": "P01", "category": "Leadership", "question": "Who is the CEO of Wipro?", "expected_entities": ["WIPRO"], "ground_truth": None},
    {"id": "P02", "category": "Leadership", "question": "Who leads Maruti Suzuki?", "expected_entities": ["MARUTI"], "ground_truth": None},
    {"id": "SH01", "category": "Shareholding", "question": "What percentage of ONGC is held by institutions?", "expected_entities": ["ONGC"], "ground_truth": None},
    {"id": "SH02", "category": "Shareholding", "question": "What is the insider holding percentage for Sun Pharmaceutical?", "expected_entities": ["SUNPHARMA"], "ground_truth": None},
    {"id": "R01", "category": "Robustness", "question": "What is the current gold price?", "expected_entities": [], "ground_truth": "Bot should state it does not have this data."},
    {"id": "R02", "category": "Robustness", "question": "Summarise the last 5 years of NIFTY performance.", "expected_entities": [], "ground_truth": None},
    {"id": "MT01", "category": "Multi-turn Sim", "question": "Tell me about TCS revenue, then compare it with Infosys revenue.", "expected_entities": ["TCS", "INFY"], "ground_truth": None},
]

# Groq-backed model names — used to decide whether to throttle/retry on 429s.
GROQ_BACKED_MODELS = {"qwen", "openai/gpt-oss"}

# ─────────────────────────────────────────────────────────────────────────────
# DATA CLASSES
# ─────────────────────────────────────────────────────────────────────────────

@dataclass
class EvalResult:
    id: str
    category: str
    question: str
    expected_entities: list
    answer: str = ""
    retrieved_context: str = ""
    latency_s: float = 0.0
    error: Optional[str] = None

    # LLM-judge scores (1-5) — left as None (not 0) whenever judging was skipped
    faithfulness: Optional[float] = None
    answer_relevance: Optional[float] = None
    completeness: Optional[float] = None

    # Derived
    avg_score: Optional[float] = None
    context_hit: Optional[bool] = None        # did the RETRIEVER surface the entity?
    answer_mentions_entity: Optional[bool] = None  # does the ANSWER text mention it?
    all_entities_hit: Optional[bool] = None   # for multi-entity (comparison) questions
    judge_reasoning: str = ""


# ─────────────────────────────────────────────────────────────────────────────
# JUDGE PROMPT
# ─────────────────────────────────────────────────────────────────────────────

JUDGE_SYSTEM = """You are an objective evaluator for a financial RAG chatbot focused on NIFTY 50 data.
Score each dimension on a scale of 1 (very poor) to 5 (excellent).
Return ONLY valid JSON. The reasoning value must be a short plain string with NO commas and NO special characters."""

JUDGE_USER = """Question: {question}

Bot Answer: {answer}

Ground Truth (if available): {ground_truth}

Score these dimensions:
1. faithfulness      - Is the answer grounded in facts? Does it avoid hallucinations?
   (5=fully grounded, 1=fabricated/contradictory)
2. answer_relevance  - Does the answer directly address what was asked?
   (5=directly answers, 1=completely off-topic)
3. completeness      - Are all parts of the question covered?
   (5=fully complete, 1=major omissions)

Return ONLY this JSON with integer scores and a short reasoning string (no commas inside the string):
{{
  "faithfulness": <1-5>,
  "answer_relevance": <1-5>,
  "completeness": <1-5>,
  "reasoning": "<ten words max no commas>"
}}"""

# ─────────────────────────────────────────────────────────────────────────────
# CORE FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def build_judge_client() -> OpenAI:
    """Build OpenAI-compatible judge. Uses Ollama locally by default."""
    base_url = os.getenv("JUDGE_BASE_URL", os.getenv("OLLAMA_BASE_URL", "http://127.0.0.1:11434/v1"))
    api_key  = os.getenv("JUDGE_API_KEY", "ollama")
    return OpenAI(base_url=base_url, api_key=api_key)


def judge_answer(
    client: OpenAI,
    question: str,
    answer: str,
    ground_truth: Optional[str],
    model: str = "llama3.2:3b",
) -> dict:
    """Call LLM judge and parse scores. Returns dict with scores or defaults on failure."""
    prompt = JUDGE_USER.format(
        question=question,
        answer=answer if answer else "(no answer - error occurred)",
        ground_truth=ground_truth or "N/A",
    )
    try:
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {"role": "system", "content": JUDGE_SYSTEM},
                {"role": "user",   "content": prompt},
            ],
            temperature=0,
            max_tokens=200,
        )
        raw = resp.choices[0].message.content.strip()
        raw = re.sub(r"```json|```", "", raw).strip()

        try:
            return json.loads(raw)
        except json.JSONDecodeError:
            pass

        def extract_score(key: str) -> Optional[int]:
            m = re.search(rf'"{key}"\s*:\s*([1-5])', raw)
            return int(m.group(1)) if m else None

        reasoning_match = re.search(r'"reasoning"\s*:\s*"([^"]*)"', raw)
        reasoning = reasoning_match.group(1) if reasoning_match else "parse error - scores extracted via regex"

        scores = {
            "faithfulness":     extract_score("faithfulness"),
            "answer_relevance": extract_score("answer_relevance"),
            "completeness":     extract_score("completeness"),
            "reasoning":        reasoning,
        }
        if any(v is not None for k, v in scores.items() if k != "reasoning"):
            print(f"  [Judge] Used regex fallback parser")
            return scores

        raise ValueError(f"Could not extract scores from: {raw[:120]}")
    except Exception as e:
        print(f"  [Judge] Failed: {e}")
        return {
            "faithfulness": None,
            "answer_relevance": None,
            "completeness": None,
            "reasoning": f"Judge error: {e}",
        }


def check_context_hit(retrieved_context: str, expected_entities: list) -> Optional[bool]:
    """
    Real retrieval-quality metric: did the RETRIEVER's returned documents
    actually contain the expected entity, regardless of what the LLM did
    with that context afterward?
    """
    if not expected_entities:
        return None
    ctx_upper = retrieved_context.upper()
    return any(e.upper() in ctx_upper for e in expected_entities)


def check_answer_mentions_entity(answer: str, expected_entities: list) -> Optional[bool]:
    """
    Generation-behavior metric (previously mislabeled 'anchor_hit'): does the
    final answer TEXT mention the expected entity? This reflects whether the
    model chose to ground its answer using available context, not whether
    retrieval succeeded — a model can retrieve the right doc and still hedge.
    """
    if not expected_entities:
        return None
    ans_upper = answer.upper()
    return any(e.upper() in ans_upper for e in expected_entities)


def check_all_entities_hit(answer: str, expected_entities: list) -> Optional[bool]:
    """Stricter version for comparison questions: were ALL entities addressed,
    not just one? The old metric used `any(...)`, which silently counted a
    comparison answer that only discussed one of the two companies as a hit."""
    if not expected_entities or len(expected_entities) < 2:
        return None
    ans_upper = answer.upper()
    return all(e.upper() in ans_upper for e in expected_entities)


_active_chain = None
_last_retrieved_docs: list = []  # populated by the retriever patch below, read after each call


def build_local_chain(model_name: str):
    import os
    import sys
    from pathlib import Path

    backend_path = Path(__file__).resolve().parent
    if str(backend_path) not in sys.path:
        sys.path.append(str(backend_path))

    from main import _build_fallback_llm, HybridNiftyRetriever, _DEFAULT_GRAPH, _DEFAULT_CHROMA
    from langchain_core.prompts import ChatPromptTemplate, MessagesPlaceholder
    from langchain_classic.chains import create_history_aware_retriever, create_retrieval_chain
    from langchain_classic.chains.combine_documents import create_stuff_documents_chain
    from langchain_groq import ChatGroq
    from langchain_openai import ChatOpenAI

    graph_path  = os.getenv("NETWORKX_GRAPH_PATH", _DEFAULT_GRAPH)
    chroma_path = os.getenv("CHROMA_PATH", _DEFAULT_CHROMA)

    if not Path(graph_path).exists():
        raise FileNotFoundError(f"Graph missing at {graph_path}")

    groq_api_key = os.getenv("GROQ_API_KEY")
    ollama_base_url = os.getenv("OLLAMA_BASE_URL", "http://127.0.0.1:11434/v1")
    ollama_model = os.getenv("OLLAMA_MODEL", "llama3.2:3b")

    # NOTE: fast_llm (query rewriter / entity extraction) is intentionally held
    # constant across every model comparison here, matching the production plan
    # (fast path always gpt-oss-20b; only the QA-generation model varies below).
    # This means retrieval/anchor-resolution quality is NOT a variable between
    # runs — differences you see in context_hit across models should be small
    # and mostly noise. If context_hit varies a lot between runs, that's a sign
    # something about the run itself (rate limiting, truncation) is interfering,
    # not that one model "retrieves better" than another — retrieval doesn't
    # depend on the QA model at all in this setup.
    fast_llm = _build_fallback_llm(
        model_id_groq="openai/gpt-oss-20b",
        model_id_ollama=ollama_model
    )

    if model_name == "qwen":
        if not groq_api_key:
            raise ValueError("GROQ_API_KEY is required to evaluate Qwen")
        qa_llm = ChatGroq(api_key=groq_api_key, model="qwen/qwen3.6-27b", temperature=0, timeout=15.0)
    elif model_name == "openai/gpt-oss":
        if not groq_api_key:
            raise ValueError("GROQ_API_KEY is required to evaluate GPT-OSS")
        qa_llm = ChatGroq(api_key=groq_api_key, model="openai/gpt-oss-20b", temperature=0, timeout=15.0)
    elif model_name == "ollama":
        qa_llm = ChatOpenAI(base_url=ollama_base_url, api_key="ollama", model=ollama_model, temperature=0, timeout=15.0)
    else:
        raise ValueError(f"Unknown model name: {model_name}")

    retriever = HybridNiftyRetriever(graph_path=graph_path, chroma_path=chroma_path, llm=fast_llm)

    # Patch: truncate context to stay under Groq's TPM cap, AND capture the
    # retrieved docs so we can compute a real context_hit metric afterward.
    original_get_relevant = retriever._get_relevant_documents

    def patched_get_relevant(query: str, **kwargs):
        global _last_retrieved_docs
        docs = original_get_relevant(query, **kwargs)
        for doc in docs:
            if len(doc.page_content) > 1000:
                doc.page_content = doc.page_content[:1000] + "\n...(truncated for evaluation)"
        docs = docs[:4]
        _last_retrieved_docs = docs  # stash for the eval loop to read
        return docs

    retriever._get_relevant_documents = patched_get_relevant

    contextualise_prompt = ChatPromptTemplate.from_messages([
        ("system", "Given a chat history and the latest user question, formulate a standalone "
                   "query that can be understood without the chat history. Do NOT answer the question."),
        MessagesPlaceholder("chat_history"),
        ("human", "{input}"),
    ])

    qa_prompt = ChatPromptTemplate.from_messages([
        ("system", """### IDENTITY
You are FinBot, a high-precision NIFTY 50 analyst. Today's date is May 2026.

### SESSION SUMMARY (prior conversations)
{session_summary}

### STRICT GROUNDING RULES
1. **Context Only**: Use ONLY the provided context. If the answer isn't there, say: "I do not have that data in my records."
2. **Priority**:
   - Use 'Source: Financial Graph' for hard numbers (Revenue, PAT, Debt).
   - Use 'Knowledge Base' for general facts or definitions.
3. **Currency**: Always use '₹' prefix. Format numbers as Cr (Crore) or Bn (Billion).
4. **Recency**: Always prioritize the most recent date found in the context.
5. **No Thought Leakage**: Do NOT include raw thinking steps, logic processes, or `<think></think>` blocks in your response. Output only the final clean answer.

### CONTEXT:
{context}
"""),
        MessagesPlaceholder("chat_history"),
        ("human", "{input}"),
    ])

    history_retriever = create_history_aware_retriever(fast_llm, retriever, contextualise_prompt)
    rag_chain = create_retrieval_chain(history_retriever, create_stuff_documents_chain(qa_llm, qa_prompt))
    return rag_chain


async def call_chat_endpoint_local(question: str, model_name: str, max_retries: int = 3) -> tuple[str, str, float]:
    """
    Invoke the chain for a single question. Returns (answer, retrieved_context, latency_s).
    Retries on Groq 429 rate-limit errors with exponential backoff instead of failing outright.
    """
    global _active_chain, _last_retrieved_docs

    for attempt in range(max_retries):
        t0 = time.perf_counter()
        try:
            res = await _active_chain.ainvoke({
                "input": question,
                "chat_history": [],
                "session_summary": ""
            })
            latency = time.perf_counter() - t0

            answer = res.get("answer")
            if answer is None:
                # Previously this crashed the regex step below with a silent
                # TypeError swallowed by the caller's except block, producing
                # confusing None-valued rows in the Excel output with no clear
                # error message. Now it's an explicit, loggable failure.
                raise ValueError(
                    "Chain returned no answer content (LLM response had empty/None "
                    "'answer' — check whether the model's tool-calling/structured "
                    "output format is compatible with this LangChain integration)"
                )

            answer = re.sub(r'<think>.*?</think>', '', answer, flags=re.DOTALL).strip()

            retrieved_context = "\n---\n".join(
                d.page_content for d in _last_retrieved_docs
            ) if _last_retrieved_docs else ""

            return answer, retrieved_context, latency

        except Exception as e:
            is_rate_limit = "429" in str(e) or "rate_limit" in str(e).lower()
            if is_rate_limit and attempt < max_retries - 1:
                backoff = 2.0 * (attempt + 1)
                print(f"    [RateLimit] Attempt {attempt+1}/{max_retries} hit 429, "
                      f"backing off {backoff:.1f}s before retry...")
                await asyncio.sleep(backoff)
                continue
            raise

    raise RuntimeError("Exhausted retries without success (unreachable)")


async def run_evaluations(
    dataset: list[dict],
    judge_model: str,
    model_name: str,
    concurrency: int,
    throttle_s: float,
) -> list[EvalResult]:
    """Run all eval questions against the chatbot locally and judge the answers."""
    judge = build_judge_client()
    results: list[EvalResult] = []
    sem = asyncio.Semaphore(concurrency)

    # Only throttle Groq-backed models — Ollama is local and has no rate limit,
    # so there's no reason to slow down local-only eval runs.
    effective_throttle = throttle_s if model_name in GROQ_BACKED_MODELS else 0.0

    async def _eval_one(item: dict) -> EvalResult:
        result = EvalResult(
            id=item["id"],
            category=item["category"],
            question=item["question"],
            expected_entities=item.get("expected_entities", []),
        )

        async with sem:
            try:
                if effective_throttle:
                    await asyncio.sleep(effective_throttle)
                answer, retrieved_context, latency = await call_chat_endpoint_local(
                    item["question"], model_name
                )
                result.answer            = answer
                result.retrieved_context = retrieved_context
                result.latency_s         = round(latency, 3)
            except Exception as e:
                result.error     = f"{type(e).__name__}: {e}"
                result.answer    = ""
                result.latency_s = 0.0
                print(f"  [{item['id']}] Local Chat error: {type(e).__name__}: {e}")

        # ── LLM Judge ──────────────────────────────────────────────────────
        # Skip judging entirely on upstream failure — previously a failed call
        # still got scored against a placeholder answer, which could produce a
        # low-but-real-looking score (e.g. 0) that's indistinguishable in the
        # spreadsheet from a genuinely bad model response. Now failures are
        # unambiguous: no scores at all, error text visible in its own column.
        if result.error:
            result.faithfulness = result.answer_relevance = result.completeness = None
            result.avg_score = None
            result.judge_reasoning = "Skipped: upstream call failed (see Error column)"
        else:
            scores = judge_answer(
                judge, item["question"], result.answer, item.get("ground_truth"), model=judge_model,
            )
            result.faithfulness     = scores.get("faithfulness")
            result.answer_relevance = scores.get("answer_relevance")
            result.completeness     = scores.get("completeness")
            result.judge_reasoning  = scores.get("reasoning", "")

            valid_scores = [s for s in [result.faithfulness, result.answer_relevance, result.completeness] if s is not None]
            result.avg_score = round(sum(valid_scores) / len(valid_scores), 2) if valid_scores else None

        result.context_hit             = check_context_hit(result.retrieved_context, result.expected_entities)
        result.answer_mentions_entity  = check_answer_mentions_entity(result.answer, result.expected_entities)
        result.all_entities_hit        = check_all_entities_hit(result.answer, result.expected_entities)

        flag = "OK" if not result.error else "FAIL"
        avg_s  = f"{result.avg_score:.2f}" if result.avg_score is not None else "N/A"
        print(f"  [{item['id']}] {flag}  avg={avg_s}  latency={result.latency_s}s")
        return result

    tasks = [_eval_one(item) for item in dataset]
    results = await asyncio.gather(*tasks)
    return list(results)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL EXPORT
# ─────────────────────────────────────────────────────────────────────────────

C_HEADER_FILL  = "1F3864"
C_HEADER_FONT  = "FFFFFF"
C_ALT_ROW      = "EBF3FF"
C_WHITE        = "FFFFFF"
C_GOOD         = "C6EFCE"
C_MID          = "FFEB9C"
C_BAD          = "FFC7CE"
C_SECTION_FILL = "D6E4F0"


def _header_style(cell, bold=True):
    cell.font      = Font(name="Arial", bold=bold, color=C_HEADER_FONT, size=10)
    cell.fill      = PatternFill("solid", fgColor=C_HEADER_FILL)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _score_fill(score: Optional[float]) -> str:
    if score is None:
        return C_WHITE
    if score >= 4.0:
        return C_GOOD
    if score >= 3.0:
        return C_MID
    return C_BAD


def _thin_border() -> Border:
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def _col_widths(ws, widths: dict):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w


_FORMULA_TRIGGER_CHARS = ("=", "+", "-", "@")


def _excel_safe(value):
    """
    Prevent Excel/openpyxl from misinterpreting text as a formula.
    Graph documents from retriever.py are formatted like '=== TCS | Company | ... ==='
    -- any cell value starting with =, +, -, or @ gets treated as a formula on save.
    Since these workbooks are never opened in real Excel to compute a result, reloading
    them (e.g. with data_only=True) silently returns None for every such cell -- this is
    exactly why "Retrieved Context" showed blank for nearly every graph-grounded answer.
    Prefixing with a leading apostrophe forces plain-text storage instead.
    """
    if isinstance(value, str) and value.startswith(_FORMULA_TRIGGER_CHARS):
        return "'" + value
    return value


def build_excel_comparison(results_dict: dict[str, list[EvalResult]], out_path: str):
    wb = Workbook()

    ws_comp = wb.active
    ws_comp.title = "Comparison Summary"
    ws_comp.sheet_view.showGridLines = False

    ws_comp.merge_cells("B1:I1")
    title_cell = ws_comp["B1"]
    title_cell.value     = "NIFTY RAG Chatbot - Model Comparison"
    title_cell.font      = Font(name="Arial", size=14, bold=True, color=C_HEADER_FONT)
    title_cell.fill      = PatternFill("solid", fgColor=C_HEADER_FILL)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_comp.row_dimensions[1].height = 32

    comp_headers = [
        "Model", "Avg Score\n(1-5)", "Faithfulness\n(1-5)",
        "Answer Relevance\n(1-5)", "Completeness\n(1-5)",
        "Avg Latency\n(s)", "Context Hit\nRate %", "Answer Mentions\nEntity %", "Error Rate\n%"
    ]

    border = _thin_border()

    ws_comp.row_dimensions[3].height = 28
    for col_idx, header in enumerate(comp_headers, start=2):
        cell = ws_comp.cell(row=3, column=col_idx, value=header)
        _header_style(cell)
        cell.border = border

    def avg(lst): return round(sum(lst)/len(lst), 2) if lst else None

    for row_idx, (model_name, results) in enumerate(results_dict.items(), start=4):
        all_faith   = [r.faithfulness     for r in results if r.faithfulness     is not None]
        all_rel     = [r.answer_relevance for r in results if r.answer_relevance is not None]
        all_comp    = [r.completeness     for r in results if r.completeness     is not None]
        all_lat     = [r.latency_s        for r in results if r.latency_s]
        all_avgs    = [r.avg_score        for r in results if r.avg_score        is not None]
        errors      = [r for r in results if r.error]

        ctx_q       = [r for r in results if r.context_hit is not None]
        ctx_hit     = [r for r in ctx_q if r.context_hit]
        ans_q       = [r for r in results if r.answer_mentions_entity is not None]
        ans_hit     = [r for r in ans_q if r.answer_mentions_entity]

        avg_score = avg(all_avgs) or 0
        f = avg(all_faith) or 0
        rel = avg(all_rel) or 0
        com = avg(all_comp) or 0
        lat = avg(all_lat) or 0
        context_hit_rate = round(len(ctx_hit)/len(ctx_q)*100, 1) if ctx_q else 0
        answer_hit_rate  = round(len(ans_hit)/len(ans_q)*100, 1) if ans_q else 0
        error_rate = round(len(errors)/len(results)*100, 1)

        row_fill = C_ALT_ROW if row_idx % 2 == 0 else C_WHITE
        ws_comp.row_dimensions[row_idx].height = 24

        data = [model_name, avg_score, f, rel, com, lat, f"{context_hit_rate}%", f"{answer_hit_rate}%", f"{error_rate}%"]
        for col_idx, val in enumerate(data, start=2):
            cell = ws_comp.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            cell.font = Font(name="Arial", size=10, bold=(col_idx == 2 or col_idx == 3))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if col_idx in (3, 4, 5, 6):
                cell.fill = PatternFill("solid", fgColor=_score_fill(val if isinstance(val, (int, float)) else None))
            else:
                cell.fill = PatternFill("solid", fgColor=row_fill)

    ws_comp.column_dimensions["B"].width = 20
    ws_comp.column_dimensions["C"].width = 15
    ws_comp.column_dimensions["D"].width = 16
    ws_comp.column_dimensions["E"].width = 18
    ws_comp.column_dimensions["F"].width = 16
    ws_comp.column_dimensions["G"].width = 16
    ws_comp.column_dimensions["H"].width = 18
    ws_comp.column_dimensions["I"].width = 14

    for model_name, results in results_dict.items():
        ws = wb.create_sheet(f"{model_name.replace('/', '_')} Details")

        headers = [
            "ID", "Category", "Question", "Answer", "Retrieved Context",
            "Faithfulness\n(1-5)", "Answer Relevance\n(1-5)", "Completeness\n(1-5)",
            "Avg Score", "Context Hit", "Answer Mentions\nEntity", "All Entities\nHit",
            "Latency (s)", "Error", "Judge Reasoning",
        ]
        ws.append(headers)
        ws.row_dimensions[1].height = 36
        for cell in ws[1]:
            _header_style(cell)

        def _bool_str(v):
            return "Yes" if v is True else ("No" if v is False else "N/A")

        for i, r in enumerate(results, start=2):
            row_fill = C_ALT_ROW if i % 2 == 0 else C_WHITE
            context_display = (r.retrieved_context[:500] + "...") if len(r.retrieved_context) > 500 else r.retrieved_context
            row_data = [
                _excel_safe(r.id), r.category, r.question, _excel_safe(r.answer),
                _excel_safe(context_display),
                r.faithfulness, r.answer_relevance, r.completeness, r.avg_score,
                _bool_str(r.context_hit), _bool_str(r.answer_mentions_entity), _bool_str(r.all_entities_hit),
                r.latency_s, r.error or "", r.judge_reasoning,
            ]
            ws.append(row_data)
            for j, cell in enumerate(ws[i]):
                cell.border    = border
                cell.font      = Font(name="Arial", size=9)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                if j in (6, 7, 8, 9):
                    cell.fill = PatternFill("solid", fgColor=_score_fill(cell.value))
                else:
                    cell.fill = PatternFill("solid", fgColor=row_fill)
                if j == 14 and cell.value:
                    cell.font = Font(name="Arial", size=9, color="C00000")

        _col_widths(ws, {
            "A": 8, "B": 18, "C": 38, "D": 45, "E": 45,
            "F": 12, "G": 12, "H": 12, "I": 10,
            "J": 11, "K": 13, "L": 11,
            "M": 10, "N": 22, "O": 40,
        })
        ws.freeze_panes = "D2"

    wb.save(out_path)
    print(f"\nComparative Excel exported to: {out_path}")


# ─────────────────────────────────────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(description="Evaluate NIFTY Graph RAG chatbot")
    p.add_argument("--dataset",     default=None, help="Path to custom JSON dataset file")
    p.add_argument("--out",         default="rag_eval_comparison.xlsx", help="Output Excel file path")
    p.add_argument("--judge-model", default="llama3.2:3b", help="LLM model for judge")
    p.add_argument("--concurrency", default=1, type=int, help="Number of parallel requests")
    p.add_argument("--throttle-s",  default=2.0, type=float,
                   help="Seconds to wait between requests for Groq-backed models only "
                        "(ollama runs are never throttled). Lower this if you're not "
                        "hitting rate limits; raise it if you are.")
    return p.parse_args()


async def main():
    import sys
    args = parse_args()

    if args.dataset:
        with open(args.dataset) as f:
            dataset = json.load(f)
        print(f"Loaded {len(dataset)} questions from {args.dataset}")
    else:
        dataset = DEFAULT_DATASET
        print(f"Using built-in dataset ({len(dataset)} questions)")

    print("\nChecking local knowledge graph resource...")
    from main import _DEFAULT_GRAPH
    graph_path = os.getenv("NETWORKX_GRAPH_PATH", _DEFAULT_GRAPH)
    if os.path.exists(graph_path):
        print("Local knowledge graph found.\n")
    else:
        print(f"Knowledge graph missing at {graph_path}\n")
        sys.exit(1)

    eval_models = ["qwen", "openai/gpt-oss", "ollama"]
    results_dict = {}

    for model in eval_models:
        print("\n" + "="*70)
        print(f"Initializing RAG chain locally for model: {model}...")
        print("="*70)

        global _active_chain
        try:
            _active_chain = build_local_chain(model)
        except Exception as e:
            print(f"Failed to build RAG chain for model '{model}': {e}")
            continue

        print(f"Running {len(dataset)} evaluations for {model} (concurrency={args.concurrency}, "
              f"throttle={args.throttle_s if model in GROQ_BACKED_MODELS else 0}s)...\n")
        results = await run_evaluations(
            dataset     = dataset,
            judge_model = args.judge_model,
            model_name  = model,
            concurrency = args.concurrency,
            throttle_s  = args.throttle_s,
        )
        results_dict[model] = results

    print("\n" + "="*75)
    print(" RAG EVALUATION COMPARATIVE SUMMARY REPORT")
    print("="*75)
    print(f"{'Model':<18} | {'Avg Score':<10} | {'Faithfulness':<12} | {'Relevance':<10} | {'Completeness':<12} | {'Avg Lat':<7}")
    print("-"*75)
    for model_name, results in results_dict.items():
        all_faith   = [r.faithfulness     for r in results if r.faithfulness     is not None]
        all_rel     = [r.answer_relevance for r in results if r.answer_relevance is not None]
        all_comp    = [r.completeness     for r in results if r.completeness     is not None]
        all_lat     = [r.latency_s        for r in results if r.latency_s]
        all_avgs    = [r.avg_score        for r in results if r.avg_score        is not None]

        def avg(lst): return round(sum(lst)/len(lst), 2) if lst else 0.00

        print(f"{model_name:<18} | {avg(all_avgs):<10.2f} | {avg(all_faith):<12.2f} | {avg(all_rel):<10.2f} | {avg(all_comp):<12.2f} | {avg(all_lat):<7.2f}s")
    print("="*75)

    build_excel_comparison(results_dict, args.out)


if __name__ == "__main__":
    asyncio.run(main())